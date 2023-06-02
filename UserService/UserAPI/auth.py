from werkzeug.security import check_password_hash, generate_password_hash
from sqlalchemy import select, insert, update
from sqlalchemy import delete as remove

from openpyxl import load_workbook
from magic import Magic
from io import BytesIO
import json

from flask import request
from flask_restx import (
    Resource,
    fields,
    Namespace,
)
from flask_jwt_extended import (
    create_access_token,
    create_refresh_token,
    jwt_required,
    get_jwt,
    get_jwt_identity
)

from service import Service
from utils import (
    retrieve_jwt,
    serialize,
    protected,
    admin_only,
    create_register_confirmation_email,
)
from config import ORM, api_config, SENDER

"""
This code creates a Flask-RESTX namespace object named "AUTH" for handling authentication-related
API endpoints. It includes a name and description for the namespace.
"""
# namespace for "/auth"
AUTH = Namespace(
    name="auth",
    description="사용자 인증을 위한 API",
)

# # required fields in request body?
# user_fields = auth.model('User', {
#     'username': fields.String(description='a User Name', required=True, example="justkode"),
#     'password': fields.String(description='Password', required=True, example="password")
# })

# # required fields in request body?
# jwt_fields = auth.model('JWT', {
#     'Authorization': fields.String(description='Authorization which you must inclued in header', required=True, example="eyJ0e~~~~~~~~~")
# })

include = ["id", "name", "type", "email", "no_show"]
exclude = ["password", "created_at"]


@AUTH.route('/register')
class Register(Service, Resource):
    """
    This code defines a Flask-RESTX endpoint for registering new users. It uses the `Service` class and
    the `ORM` configuration to interact with the database. The `post` method handles the HTTP POST
    request and validates the request body arguments using the `User` model. It then checks if the user
    already exists in the database and returns an error message if so. If the user does not exist, it
    generates a password hash and creates a new user in the database using the `insert` method. Finally,
    it returns the newly created user as a JSON response.
    """
    def __init__(self, *args, **kwargs):
        """
        This is the initialization function for a class that inherits from both Service and Resource
        classes, passing in arguments and keyword arguments.
        """
        Service.__init__(
            self, model_config=ORM, api_config=api_config
        )
        Resource.__init__(self, *args, **kwargs)

    @jwt_required()
    @admin_only()
    def admin_credentials_required(*args, **kwargs):
        pass

    def send_email(self, user):
        # create and send email object
        email_object = create_register_confirmation_email(user, sender=SENDER)
        
        try:
            email_resp = self.query_api(
                "send_email", "post",
                headers=request.headers, body=json.dumps(email_object)
            )
            
        except:
            return None

        return email_resp

    def post(self):
        """
        This is a Python function for registering new users with validation and password hashing.

        :return: a JSON response with a status code of 200 or 500 depending on whether the registration
        was successful or not. If successful, the response includes a serialized representation of the
        newly created user. If unsuccessful, the response includes an error message.

        ---
        request body:
        - username: new username
        - password: new password (currently has no length checks etc)
        """

        
        # using orm model
        try:
            with self.query_model("User") as (conn, User):
                # validate request body arguments
                req, invalidated = User.validate(request.json)
                if len(invalidated) != 0:
                    return {
                        "status": False,
                        "msg": "key:value pair wrong",
                        "invalid": invalidated
                    }, 200

                # check if user exists
                res = conn.execute(
                    select(User).where(User.id == req["id"])
                ).mappings().all()

                if len(res) != 0:
                    return {
                        "status": False,
                        "msg": "User Exists"
                    }, 200

                # hash password
                req["password"] = generate_password_hash(req["password"])
                
                # create new user
                conn.execute(
                    insert(User), {
                        **req
                    }
                )

                # return new user
                res = conn.execute(
                    select(User).where(User.id == req["id"])
                ).mappings().fetchone()

                # send email to new user
                _ = self.send_email(serialize(res, exclude=exclude))

                return {
                    "status": True,
                    "msg": "success",
                    "User": serialize(res, exclude=exclude)
                }, 200

        except Exception as e:
            print(e, flush=True)
            return {
                "status": False,
                "msg": "Register Failed",
            }, 500


@AUTH.route('/login')
class Login(Service, Resource):
    """
    This code defines a Flask-RESTX endpoint for logging in users. It uses the `Service` class and the
    `ORM` configuration to interact with the database. The `post` method handles the HTTP POST request
    and validates the request body arguments using the `User` model. It then checks if the user exists
    in the database and returns an error message if not. If the user exists, it checks if the password
    is correct using the `check_password_hash` function from the `werkzeug.security` module. If the
    password is correct, it generates an access token and a refresh token using the
    `create_access_token` and `create_refresh_token` functions from the `flask_jwt_extended` module.
    Finally, it returns the access token, refresh token, and user information as a JSON response.
    """
    def __init__(self, *args, **kwargs):
        """
        This is the initialization function for a class that inherits from both Service and Resource
        classes, passing ORM as a model configuration parameter to the Service class.
        """
        Service.__init__(self, model_config=ORM)
        Resource.__init__(self, *args, **kwargs)

    def post(self):
        """
        This is a Python function that handles user login by validating the request body, checking if
        the user exists and if the password is correct, and then creating access and refresh tokens.
        :return: a dictionary with the following keys and values:
        - "status": a boolean value indicating whether the login was successful or not
        - "access_token": a string representing the access token generated for the user
        - "refresh_token": a string representing the refresh token generated for the user
        - "User": a dictionary representing the user object, with keys and values based on the exclude
        and include.

        ---
        req body:
        - username: existing username
        - password: plaintext password
        """
        
        # using orm model
        try:
            with self.query_model("User") as (conn, User):
                # validate request body
                req, invalidated = User.validate(request.json, optional=True)
                print(req, invalidated)
                if len(invalidated) != 0:
                    return {
                        "status": False,
                        "msg": "key:value pair wrong",
                        "invalid": invalidated
                    }, 200

                # get user
                res = conn.execute(
                    select(User).where(User.id == req["id"])
                ).mappings().fetchone()

                # check if user exists
                if res is None:
                    return {
                        "status": False,
                        "msg": "User Not Found"
                    }, 200

                # check if password is right
                if not check_password_hash(res["password"], req["password"]):
                    return {
                        "status": False,
                        "msg": "Wrong Password"
                    }, 200

                # TODO: replace names with underscore names
                identity = serialize(res, include=include)
                access_token = create_access_token(
                    identity=identity,
                    fresh=True
                )
                refresh_token = create_refresh_token(
                    identity=identity
                )
                
                return {
                    "status": True,
                    "access_token": access_token,
                    "refresh_token": refresh_token,
                    "User": serialize(res, exclude=exclude)
                }, 200

        except Exception as e:
            print(e, flush=True)
            return {
                "status": False,
                "msg": "Login Failed"
            }, 500

@AUTH.route('/logout')
class Logout(Service, Resource):
    """
    The above code defines a Flask route for logging out a user. It uses the Flask-JWT-Extended library
    to require a valid access token for the route to be accessed. When the route is accessed, it adds
    the token to a token blocklist to prevent it from being used again, and returns a response
    indicating whether the token was successfully revoked or had already been revoked.
    """
    def __init__(self, *args, **kwargs):
        """
        This is the initialization function for a class that inherits from both Service and Resource
        classes, passing ORM as a model configuration parameter to the Service class.
        """
        Service.__init__(self, model_config=ORM)
        Resource.__init__(self, *args, **kwargs)

    
    def add_token_to_blocklist(self, token):
        """
        The function adds a token to a blocklist if it is not already present.
        
        :param token: The `token` parameter is a dictionary that represents a JSON Web Token (JWT) that
        needs to be added to a blocklist.
        
        :return: a boolean value. It returns True if the token was successfully added to the blocklist,
        and False if the token was already in the blocklist.
        """
        # from config import TOKEN_BLOCKLIST
        
        # if jti not in TOKEN_BLOCKLIST:
        #     TOKEN_BLOCKLIST.add(jti)
        #     return True
        # return False
        with self.query_model("Token_Blocklist") as (conn, Blocklist):
            res = conn.execute(
                select(Blocklist).where(Blocklist.jti == token["jti"])
            ).mappings().fetchone()
            if res is not None:
                return False

            res = conn.execute(
                insert(Blocklist),
                {
                    "jti": token["jti"],
                    "user_id": token["sub"]["id"],
                    "type": token["sub"]["type"]
                }
            )
            return True

    @jwt_required(verify_type=False)
    def delete(self):
        """
        This function logs out a user by revoking their JWT token.
        
        :return: The `delete` method is returning a dictionary with a "status" key and a "msg" key. The
        value of the "status" key is a boolean indicating whether the token was successfully revoked or
        not. The value of the "msg" key is a string message indicating whether the token was already
        revoked or has been successfully revoked.
        """
        jwt = get_jwt()
        if self.add_token_to_blocklist(jwt):
            return {
                "status": True,
                "msg": "token revoked"
            }
        return {
            "status": True,
            "msg": "already revoked"
        }

@AUTH.route("/jwt-status")
class JWTStatus(Service, Resource):
    """
    The above code defines a Flask endpoint for checking if a user's JWT token is valid. The endpoint
    requires a valid JWT token to be included in the request header. If the token is valid, the endpoint
    returns a JSON response with a status of True, a message of "authenticated", and the serialized user
    object. If the token is invalid or an error occurs, the endpoint returns a JSON response with a
    status of False and a message of "jwt status Failed".
    """
    def __init__(self, *args, **kwargs):
        """
        This is the initialization function for a class that inherits from both Service and Resource
        classes, passing ORM as a model configuration parameter to the Service class.
        """
        Service.__init__(self, model_config=ORM)
        Resource.__init__(self, *args, **kwargs)

    @jwt_required()
    def get(self):
        """
        This is a Python function that checks if a user's JWT token is valid and returns the user's
        information if it is.

        :return: This code returns a JSON response with a status code of 200 if the user's JWT token is
        valid and the user is authenticated. The response includes a message indicating successful
        authentication and the serialized user object. If there is an exception, the code returns a JSON
        response with a status code of 500 and a message indicating that the JWT status has failed.

        ---
        req header
        - Authentication: Bearer [JWT Token]
        """
        try:             
            with self.query_model("User") as (conn, User):
                identity = retrieve_jwt()
                res = conn.execute(
                    select(User).where(User.id == identity["id"])
                ).mappings().fetchone()

                return {
                    "status": True,
                    "msg": "authenticated",
                    "User": serialize(res, exclude=exclude)
                }, 200

        except Exception as e:
            print(e, flush=True)
            return {
                "status": False,
                "msg": "jwt status Failed"
            }, 500


@AUTH.route("/jwt-refresh")
class JWTRefresh(Service, Resource):
    """
    The above code is defining a Flask endpoint for refreshing a JWT access token. It requires a valid
    refresh token to be included in the request header, which is checked using the `jwt_required`
    decorator with the `refresh=True` parameter. If the refresh token is valid, a new access token is
    created using the `create_access_token` function from the Flask-JWT-Extended library and returned in
    the response along with a status code of 200.
    """
    def __init__(self, *args, **kwargs):
        """
        This is the initialization function for a class that inherits from both Service and Resource
        classes, passing ORM as a model configuration parameter to the Service class.
        """
        Service.__init__(self, model_config=ORM)
        Resource.__init__(self, *args, **kwargs)

    @jwt_required(refresh=True)
    def get(self):
        """
        This is a Python function that checks if a user token is valid and returns a new access token.
        
        :return: This endpoint returns a JSON object with a boolean value indicating the status of the
        request and a new access token generated using the identity retrieved from the current JWT
        token. The HTTP status code returned is 200.

        ---
        req header
        - Authentication: Bearer [JWT Token]
        """

        identity = serialize(get_jwt_identity(), include=include)
        access_token = create_access_token(identity=identity)
        return {
            "status": True,
            "access_token": access_token,
        }, 200
        
            

@AUTH.route("/import-users")
class UserImport(Service, Resource):
    """
    The above code is a Python Flask API endpoint that allows authorized admin users to import user data
    from an Excel file. The endpoint accepts a POST request with an Excel file attached, validates the
    file type, reads the file, iterates over the rows, validates the data, and inserts the valid user
    data into the database. The endpoint returns a JSON response indicating the success or failure of
    the import operation, along with the number of users imported.
    """
    allowed_filetypes = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/x-ole-storage"
    ]

    def __init__(self, *args, **kwargs):
        """
        This is the initialization function for a class that inherits from both Service and Resource
        classes, passing ORM as a model configuration parameter to the Service class.
        """
        Service.__init__(self, model_config=ORM)
        Resource.__init__(self, *args, **kwargs)

    @jwt_required()
    @admin_only()
    def post(self):        
        """
        This is a Python function that imports user data from an Excel file, validates the data, and
        inserts it into a database.
        :return: a dictionary with keys "status", "msg", and "imported_count". The values of these keys
        depend on the outcome of the function's execution. If the file is not selected or has an invalid
        filetype, the "status" key will be False and the "msg" key will contain an error message. If the
        file is successfully imported, the "status" key will
        """
        f = request.files.get('file')
        
        # If the user does not select a file, the browser submits an
        # empty file without a filename.
        if f is None or f.filename == '':
            return {
                "status": False,
                "msg": "no file selected"
            }, 200

        f = f.read()
        
        # check filetype
        mime = Magic(mime=True)
        if mime.from_buffer(f) not in self.allowed_filetypes:
            return {
                "status": False,
                "msg": "invalid filetype"
            }, 200

        try:
            with self.query_model("User") as (conn, User):
                users_sheet = load_workbook(filename=BytesIO(f)).active

                schema = []
                insert_values = []
                for i, row in enumerate(users_sheet.iter_rows(min_row=1, min_col=1, max_col=7)):
                    # set excel file column keys
                    if i == 0:
                        schema = [v.value for v in row]
                        continue
                    
                    # break if end of file
                    if not any([r.value for r in row]):
                        break

                    # get values
                    current = {}
                    for k, v in zip(schema, row):
                        # get python type of col
                        which_type = User.columns.get("User."+k).type.python_type

                        # format float to int for numbers in excel file
                        v.value = int(v.value) if type(v.value) is float else v.value
                        
                        # convert to proper type
                        current[k] = v.value if which_type is None else which_type(v.value)

                    # create new user dict
                    new_user, invalid = User.validate(
                        current,
                        optional=False,
                    )

                    print("validated: ", new_user, invalid, flush=True)

                    #check excel values
                    if len(invalid) != 0:
                        return {
                            "status": False,
                            "msg": "invalid value",
                            "invalid": invalid
                        }, 200

                    # check if user in db 
                    res = conn.execute(
                        select(User).where(User.id == new_user["id"])
                    ).mappings().fetchone()
                    if res is None:
                        new_user["password"] = generate_password_hash(new_user["password"])
                        insert_values.append(new_user)
                        print("inserting: ", new_user, flush=True)
                        
                # create new user
                if insert_values:
                    conn.execute(insert(User), insert_values)

            return {
                "status": True,
                "msg": "users imported",
                "imported_count": len(insert_values)
            }, 200

        except Exception:
        #    print(e, flush=True)
            return {
                "status": False,
                "msg": "error while importing users"
            }, 500







