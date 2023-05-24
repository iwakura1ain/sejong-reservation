from werkzeug.security import check_password_hash, generate_password_hash
from sqlalchemy import select, insert, update
from sqlalchemy import delete as remove

from openpyxl import load_workbook
from magic import Magic
from io import BytesIO

from flask import request
from flask_restx import (
    Resource,
    fields,
    Namespace,
)
from flask_jwt_extended import (
    create_access_token,
    create_refresh_token,
    jwt_required,
    get_jwt
)

from service import Service
from utils import retrieve_jwt, serialize, protected, admin_only
from config import ORM, API

"""
This code creates a Flask-RESTX namespace object named "AUTH" for handling authentication-related
API endpoints. It includes a name and description for the namespace.
"""
# namespace for "/manage"
MANAGE = Namespace(
    name="manage",
    description="사용자 위한 API",
)

@MANAGE.route("/import-users")
class UserImport(Service, Resource):
    allowed_filetypes = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/x-ole-storage"
    ]

    def __init__(self, *args, **kwargs):
        Service.__init__(self, model_config=ORM)
        Resource.__init__(self, *args, **kwargs)

    @jwt_required()
    @admin_only()
    def post(self):
        f = request.files['file']

        # If the user does not select a file, the browser submits an
        # empty file without a filename.
        if f.filename == '':
            return {
                "status": False,
                "msg": "no file selected"
            }, 200

        f = f.read()
        
        # check filetype
        mime = Magic(mime=True)
        if mime.from_buffer(f) not in self.allowed_filetypes:
            return {
                "status": False,
                "msg": "invalid filetype"
            }, 200

        with self.query_model("User") as (conn, User):
            users_sheet = load_workbook(filename=BytesIO(f)).active
            schema = User.columns

            insert_values = []
            for row in users_sheet.iter_rows(min_row=2, min_col=1, max_col=8):
                # create new user dict
                new_user, invalid = User.validate(
                    {key: val.value for key, val in zip(schema, row)}
                )

                #check excel values
                if len(invalid) != 0:
                    return {
                        "status": False,
                        "msg": "invalid value",
                        "invalid": row
                    }, 200

                # check if user in db 
                res = conn.execute(
                    select(User).where(User.id == new_user["id"])
                ).mappings().fetchone()
                if res is None:
                    new_user["password"] = generate_password_hash(new_user["password"])
                    insert_values.append(new_user)

            # insert new user
            conn.execute(
                insert(User), insert_values
            )

        return {
            "status": True,
            "msg": "users imported",
            "imported_count": len(insert_values)
        }, 200


